
# Discord bot to submit and manage leave applications

import discord
from discord.ext import commands
from discord.ui import Button, View, Modal, TextInput
import sqlite3
import re
import os
import pandas as pd
import time
import asyncio
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# Replace with your bot token
TOKEN = 'MTI4NDM4ODM5MjgzODEwMzEwMQ.Gz3UDQ.TeVOEWv2zQJJz-SP-Mu9njYyYZsPuH50c4KO5o'

# Channel ID for submitting leave applications
submit_channel_id = 1281201690904629289

# Role ID for employees
emp_role_id = 1290199089371287562

# Approval channels
approval_channels = {
    'architects': 1298229045229781052,
    'site': 1298229146228359200,
    'cad': 1298229187865346048,
    'administration': 1298229241338662932,
    'hr': 1283723426103562343,
    'pa': 1283723484698120233,
    'interns': 1298229045229781052,
    'heads': 1283723426103562343,
    'project_coordinator': 1283723426103562343
}

# Role IDs for departments
department_roles = {
    'architects': 1281172225432752149,
    'site': 1285183387258327050,
    'cad': 1281172603217645588,
    'administration': 1281171713299714059,
    'interns': 1281195640109400085,
}

# Role IDs for direct second approval
direct_second_approval_roles = {
    'project_coordinator': 1298230195991478322,
    'heads': 1281173876704804937
}

# Intents for the bot
intents = discord.Intents.default()
intents.messages = True
intents.members = True
intents.message_content = True

client = commands.Bot(command_prefix='!', intents=intents)

# Network path to the NAS server
nas_server_path = r"Z:\Bot databases\Leave"

# Initialize SQLite databases on the NAS server
conn1 = sqlite3.connect(os.path.join(nas_server_path, 'leave_details.db'))
conn2 = sqlite3.connect(os.path.join(nas_server_path, 'dynamic_updates.db'))
c1 = conn1.cursor()
c2 = conn2.cursor()

# Initialize the rate limiter
rate_limiter = asyncio.Semaphore(1)  # Adjust the number as needed

async def rate_limited_request(coro):
    async with rate_limiter:
        await asyncio.sleep(1)  # Add a delay between requests
        return await coro

# Initialize the queue
db_queue = asyncio.Queue()

async def db_worker():
    while True:
        func, args, kwargs, future = await db_queue.get()
        try:
            result = func(*args, **kwargs)
            future.set_result(result)
        except Exception as e:
            future.set_exception(e)
        db_queue.task_done()

# Function to create a leave application embed
def create_leave_embed(leave_details, user_id, nickname, current_stage):
    embed = discord.Embed(title="Leave Application", color=5810975)
    embed.add_field(name="Name", value=nickname, inline=False)
    if leave_details.get('leave_type'):
        embed.add_field(name="Leave Type", value=leave_details['leave_type'], inline=False)
    if leave_details.get('leave_reason'):
        embed.add_field(name="Leave Reason", value=leave_details['leave_reason'], inline=False)
    if leave_details.get('leave_duration'):
        embed.add_field(name="Leave Duration", value=leave_details['leave_duration'], inline=False)
    if leave_details.get('date_from'):
        embed.add_field(name="Date From", value=leave_details['date_from'], inline=False)
    if leave_details.get('date_to'):
        embed.add_field(name="Date To", value=leave_details['date_to'], inline=False)
    if leave_details.get('number_of_days_off') is not None:
        embed.add_field(name="Number of Days Off", value=leave_details['number_of_days_off'], inline=False)
    if leave_details.get('resume_office_on'):
        embed.add_field(name="Resume Office On", value=leave_details['resume_office_on'], inline=False)
    if leave_details.get('time_off'):
        embed.add_field(name="Time Off", value=leave_details['time_off'], inline=False)
    if leave_details.get('leave_id') is not None:
        embed.add_field(name="Leave ID", value=leave_details['leave_id'], inline=False)
    if leave_details.get('approved_by'):
        embed.add_field(name="Approved By", value=leave_details['approved_by'], inline=False)
    embed.set_footer(text=f"Stage: {current_stage} | User ID: {user_id}")
    return embed

# Function to sanitize table names
def sanitize_table_name(name):
    sanitized_name = re.sub(r'\W+', '_', name)
    return sanitized_name

async def db_execute(func, *args, **kwargs):
    future = asyncio.Future()
    await db_queue.put((func, args, kwargs, future))
    return await future

# Function to create a table for a user in the first database
async def create_user_table(nickname):
    table_name = sanitize_table_name(nickname)
    await db_execute(c1.execute, f'''
        CREATE TABLE IF NOT EXISTS {table_name} (
            leave_id INTEGER PRIMARY KEY AUTOINCREMENT,
            leave_type TEXT,
            leave_reason TEXT,
            date_from TEXT,
            date_to TEXT,
            number_of_days_off REAL,
            resume_office_on TEXT,
            time_off TEXT,
            leave_status TEXT,
            reason_for_decline TEXT,
            approved_by TEXT,
            time_period TEXT,
            footer_text TEXT
        )
    ''')
    await db_execute(conn1.commit)

# Function to delete a table for a user in the first database
async def delete_user_table(nickname):
    table_name = sanitize_table_name(nickname)
    await db_execute(c1.execute, f'''
        DROP TABLE IF EXISTS {table_name}
    ''')
    await db_execute(conn1.commit)

# Function to create the universal table in the second database
async def create_dynamic_table():
    await db_execute(c2.execute, '''
        CREATE TABLE IF NOT EXISTS dynamic_updates (
            nickname TEXT,
            user_id INTEGER UNIQUE,
            total_sick_leave REAL DEFAULT 0,
            total_casual_leave REAL DEFAULT 0,
            total_c_off REAL DEFAULT 0,
            last_leave_taken TEXT,
            off_duty_hours REAL DEFAULT 0
        )
    ''')
    await db_execute(conn2.commit)

# Function to export leave details to an Excel file
async def export_leave_details_to_excel():
    # Specify the directory where you want to save the Excel file
    export_directory = r"Z:\Bot databases\Leave details"
    os.makedirs(export_directory, exist_ok=True)  # Create the directory if it doesn't exist

    # Generate the file name with the current date
    current_date = datetime.now().strftime("%d-%m-%Y")
    export_path = os.path.join(export_directory, f'leave_details_export_{current_date}.xlsx')

    # Calculate the start and end dates of the current month
    start_of_month = datetime.now().replace(day=1).strftime("%d-%m-%Y")
    end_of_month = (datetime.now().replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    end_of_month = end_of_month.strftime("%d-%m-%Y")

    # Fetch all tables in the leave_details database, excluding sqlite_sequence
    tables = await db_execute(c1.execute, "SELECT name FROM sqlite_master WHERE type='table' AND name != 'sqlite_sequence';")
    tables = tables.fetchall()

    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)

    for table_name in tables:
        table_name = table_name[0]
        # Truncate the table name to 31 characters for the sheet name
        sheet_name = table_name[:31]
        # Read the table into a DataFrame, including only accepted and withdrawn leaves within the month
        query = f"""
            SELECT * FROM {table_name}
            WHERE (leave_status = 'Accepted' OR leave_status = 'Withdrawn')
            AND (date_from >= '{start_of_month}' AND (date_to <= '{end_of_month}' OR date_to IS NULL))
        """
        df = pd.read_sql_query(query, conn1)
        
        if df.empty:
            continue  # Skip empty tables
        
        # Add a new sheet to the workbook
        ws = wb.create_sheet(title=sheet_name)
        
        # Write the DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Apply header formatting
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Create a table for better layout
        tab = Table(displayName=f"Table_{sheet_name}", ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

    # Check if any sheets were added to the workbook
    if len(wb.sheetnames) == 0:
        print("No data available to export.")
        return
    else:
        # Save the workbook
        wb.save(export_path)
        print(f"Leave details exported to {export_path}")

# Create tables for all users in the emp role
@client.event
async def on_ready():
    print(f'Logged in as {client.user}')
    guild = client.guilds[0]
    emp_role = guild.get_role(emp_role_id)
    await create_dynamic_table()  # Await the coroutine
    for member in emp_role.members:
        await create_user_table(member.display_name)  # Await the coroutine
        await db_execute(c2.execute, '''
            INSERT OR IGNORE INTO dynamic_updates (nickname, user_id)
            VALUES (?, ?)
        ''', (member.display_name, member.id))
    await db_execute(conn2.commit)
    
    # Reattach view to the command button in the submit channel
    channel = client.get_channel(submit_channel_id)
    async for message in channel.history(limit=10):
        if message.author == client.user and message.embeds and "Please select your leave type using the buttons given below" in message.embeds[0].title:
            view = LeaveApplicationView()
            await rate_limited_request(message.edit(view=view))
            print("Reattached view to existing message in submit channel.")
            break
    else:
        # If no existing message is found, create a new one
        view = LeaveApplicationView()
        await rate_limited_request(channel.send(embed=discord.Embed(title="Please select your leave type using the buttons given below", color=5810975), view=view))

    # Reattach views to leave approval embeds in all approval channels
    for channel_id in approval_channels.values():
        channel = client.get_channel(channel_id)
        async for message in channel.history(limit=50):  # Adjust the limit as needed
            if message.author == client.user and message.embeds and "Leave Application" in message.embeds[0].title:
                if not any(button.disabled for button in message.components[0].children):
                    leave_details = extract_leave_details_from_embed(message.embeds[0])
                    footer_text = message.embeds[0].footer.text if message.embeds[0].footer else None
                    if footer_text:
                        # Extract user_id, current_stage, channel_id, and message_id from footer text
                        parts = footer_text.split(" | ")
                        if len(parts) >= 5:
                            try:
                                current_stage = parts[0].split(": ")[1]
                                user_id = int(parts[1].split(": ")[1])
                                nickname = parts[2].split(": ")[1]  # Extract nickname from footer text
                                channel_id = int(parts[3].split(": ")[1])  # Extract channel ID from footer text
                                message_id = int(parts[4].split(": ")[1])  # Extract message ID from footer text
                                view = LeaveApprovalView(user_id, leave_details, current_stage, nickname)
                                message = await rate_limited_request(channel.fetch_message(message_id))
                                await rate_limited_request(message.edit(view=view))
                                print(f"Reattached view to leave approval message in channel {channel_id}.")
                            except (IndexError, ValueError) as e:
                                print(f"Error parsing footer text in channel {channel_id}: {footer_text} - {e}")
                        else:
                            print(f"Skipping message in channel {channel_id} due to incorrect footer format: {footer_text}")
                    else:
                        # Retrieve footer text from the database
                        await db_execute(c1.execute, f'''
                            SELECT footer_text FROM {sanitize_table_name(nickname)}
                            WHERE leave_id = ?
                        ''', (leave_details['leave_id'],))
                        db_footer_text = c1.fetchone()
                        if db_footer_text:
                            parts = db_footer_text[0].split(" | ")
                            if len(parts) >= 5:
                                try:
                                    current_stage = parts[0].split(": ")[1]
                                    user_id = int(parts[1].split(": ")[1])
                                    nickname = parts[2].split(": ")[1]  # Extract nickname from footer text
                                    channel_id = int(parts[3].split(": ")[1])  # Extract channel ID from footer text
                                    message_id = int(parts[4].split(": ")[1])  # Extract message ID from footer text
                                    view = LeaveApprovalView(user_id, leave_details, current_stage, nickname)
                                    message = await rate_limited_request(channel.fetch_message(message_id))
                                    await rate_limited_request(message.edit(view=view))
                                    print(f"Reattached view to leave approval message in channel {channel_id}.")
                                except (IndexError, ValueError) as e:
                                    print(f"Error parsing footer text from database in channel {channel_id}: {db_footer_text[0]} - {e}")
                            else:
                                print(f"Skipping message in channel {channel_id} due to incorrect footer format in database: {db_footer_text[0]}")
                        else:
                            print(f"Skipping message in channel {channel_id} due to missing footer text in database.")

    # Start the database worker task
    client.loop.create_task(db_worker())

def extract_leave_details_from_embed(embed):
    leave_details = {}
    for field in embed.fields:
        if field.name == "Leave Type":
            leave_details['leave_type'] = field.value
        elif field.name == "Leave Reason":
            leave_details['leave_reason'] = field.value
        elif field.name == "Date From":
            leave_details['date_from'] = field.value
        elif field.name == "Date To":
            leave_details['date_to'] = field.value
        elif field.name == "Number of Days Off":
            leave_details['number_of_days_off'] = float(field.value)
        elif field.name == "Resume Office On":
            leave_details['resume_office_on'] = field.value
        elif field.name == "Time Off":
            leave_details['time_off'] = field.value
        elif field.name == "Leave ID":
            leave_details['leave_id'] = int(field.value)
        elif field.name == "Approved By":
            leave_details['approved_by'] = field.value
    return leave_details

# Create or delete table for users who join or leave the emp role
@client.event
async def on_member_update(before, after):
    if emp_role_id in [role.id for role in after.roles] and emp_role_id not in [role.id for role in before.roles]:
        create_user_table(after.display_name)
        c2.execute('''
            INSERT OR IGNORE INTO dynamic_updates (nickname, user_id)
            VALUES (?, ?)
        ''', (after.display_name, after.id))
        conn2.commit()
    elif emp_role_id not in [role.id for role in after.roles] and emp_role_id in [role.id for role in before.roles]:
        delete_user_table(after.display_name)
        c2.execute('''
            DELETE FROM dynamic_updates
            WHERE user_id = ?
        ''', (after.id,))
        conn2.commit()

class LeaveApplicationView(View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="Full Day", style=discord.ButtonStyle.primary, custom_id="FormID1")
    async def full_day_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(FullDayLeaveModal())

    @discord.ui.button(label="Half Day", style=discord.ButtonStyle.primary, custom_id="FormID2")
    async def half_day_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(HalfDayLeaveModal())

    @discord.ui.button(label="Off Duty", style=discord.ButtonStyle.primary, custom_id="FormID3")
    async def off_duty_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(OffDutyLeaveModal())

    @discord.ui.button(label="Leave Details", style=discord.ButtonStyle.secondary, custom_id="FormID4")
    async def leave_details_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        user_id = interaction.user.id
        
        # Fetch the last accepted leave date, total casual leave, and total sick leave from the dynamic_updates database
        result = await db_execute(c2.execute, '''
            SELECT last_leave_taken, total_casual_leave, total_sick_leave
            FROM dynamic_updates
            WHERE user_id = ?
        ''', (user_id,))
        result = result.fetchone()
        
        if result:
            last_leave_taken, total_casual_leave, total_sick_leave = result
            await interaction.response.send_message(
                f"Last Accepted Leave Date: {last_leave_taken}\n"
                f"Total Casual Leave: {total_casual_leave}\n"
                f"Total Sick Leave: {total_sick_leave}", 
                ephemeral=True
            )
        else:
            await interaction.response.send_message("No leave details found.", ephemeral=True)

        @discord.ui.button(label="Withdraw Leave", style=discord.ButtonStyle.danger, custom_id="FormID5")
        async def withdraw_leave_button(self, interaction: discord.Interaction, button: discord.ui.Button):
            await interaction.response.send_modal(WithdrawLeaveModal())

class LeaveApprovalView(View):
    def __init__(self, user_id, leave_details, current_stage, nickname):
        super().__init__(timeout=None)
        self.user_id = user_id
        self.leave_details = leave_details
        self.current_stage = current_stage
        self.nickname = nickname  # Store the name of the leave applier

    @discord.ui.button(label="Accept", style=discord.ButtonStyle.success, custom_id="accept_leave")
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_approval(interaction, approved=True)

    @discord.ui.button(label="Decline", style=discord.ButtonStyle.danger, custom_id="decline_leave")
    async def decline_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_approval(interaction, approved=False)

    async def handle_approval(self, interaction: discord.Interaction, approved: bool):
        try:
            if approved:
                # Store the name of the user who accepted the leave
                self.leave_details['approved_by'] = interaction.user.display_name

                # Update the database with the approver's name
                table_name = sanitize_table_name(self.nickname)
                c1.execute(f'''
                    UPDATE {table_name}
                    SET approved_by = ?
                    WHERE leave_id = ?
                ''', (interaction.user.display_name, self.leave_details['leave_id']))
                conn1.commit()

                # Create the embed without changing the leave applier's name
                embed = create_leave_embed(self.leave_details, self.user_id, self.nickname, self.current_stage)
                if 'Approved By' not in [field.name for field in embed.fields]:
                    embed.add_field(name="Approved By", value=interaction.user.display_name, inline=False)

                if self.current_stage == 'first':
                    # Determine the second approval channel
                    second_approval_channel_id = approval_channels.get('hr')  # Default second approval channel
                    for role_name, role_id in direct_second_approval_roles.items():
                        if role_id in [role.id for role in interaction.user.roles]:
                            second_approval_channel_id = approval_channels.get(role_name)
                            break

                    # Send the embed to the second approval channel
                    second_approval_channel = client.get_channel(second_approval_channel_id)
                    if second_approval_channel:
                        message = await second_approval_channel.send(embed=embed, view=LeaveApprovalView(self.user_id, self.leave_details, 'second', self.nickname))
                    else:
                        raise ValueError(f"Channel with ID {second_approval_channel_id} not found.")

                    # Update the footer text in the database
                    new_footer_text = f"Stage: second | User ID: {self.user_id} | Nickname: {self.nickname} | Channel ID: {second_approval_channel_id} | Message ID: {message.id}"
                    embed.set_footer(text=new_footer_text)
                    await message.edit(embed=embed)
                    c1.execute(f'''
                        UPDATE {table_name}
                        SET footer_text = ?
                        WHERE leave_id = ?
                    ''', (new_footer_text, self.leave_details['leave_id']))
                    conn1.commit()

                    # Update the original message to show it has been approved
                    await interaction.message.edit(view=None)
                    await interaction.message.edit(content="Leave Approved", view=View().add_item(Button(label="Approved", style=discord.ButtonStyle.success, disabled=True)))
                    await interaction.response.send_message("Leave approved and sent to the next approval stage.", ephemeral=True)

                    # Send a message to the leave applier
                    user = await client.fetch_user(self.user_id)
                    await user.send("Your leave has been recommended by your head of department to the HR.")

                elif self.current_stage == 'second':
                    # Update the leave status in the database
                    c1.execute(f'''
                        UPDATE {table_name}
                        SET leave_status = 'Accepted'
                        WHERE leave_id = ?
                    ''', (self.leave_details['leave_id'],))
                    conn1.commit()

                    # Update the dynamic_updates table
                    leave_reason = self.leave_details['leave_reason'].lower()
                    number_of_days_off = self.leave_details['number_of_days_off']
                    if leave_reason == "sick":
                        c2.execute(f'''
                            UPDATE dynamic_updates
                            SET total_sick_leave = total_sick_leave + ?, last_leave_taken = ?
                            WHERE user_id = ?
                        ''', (number_of_days_off, self.leave_details['date_to'], self.user_id))
                    elif leave_reason == "casual":
                        c2.execute(f'''
                            UPDATE dynamic_updates
                            SET total_casual_leave = total_casual_leave + ?, last_leave_taken = ?
                            WHERE user_id = ?
                        ''', (number_of_days_off, self.leave_details['date_to'], self.user_id))
                    elif leave_reason == "c. off":
                        c2.execute(f'''
                            UPDATE dynamic_updates
                            SET total_c_off = total_c_off + ?, last_leave_taken = ?
                            WHERE user_id = ?
                        ''', (number_of_days_off, self.leave_details['date_to'], self.user_id))
                    conn2.commit()

                    # Update the footer text in the database
                    new_footer_text = f"Stage: third | User ID: {self.user_id} | Nickname: {self.nickname} | Channel ID: {approval_channels['pa']} | Message ID: {interaction.message.id}"
                    embed.set_footer(text=new_footer_text)
                    await interaction.message.edit(embed=embed)
                    c1.execute(f'''
                        UPDATE {table_name}
                        SET footer_text = ?
                        WHERE leave_id = ?
                    ''', (new_footer_text, self.leave_details['leave_id']))
                    conn1.commit()

                    # Update the original message to show it has been approved
                    await interaction.message.edit(view=None)
                    await interaction.message.edit(content="Leave Approved", view=View().add_item(Button(label="Approved", style=discord.ButtonStyle.success, disabled=True)))
                    await interaction.response.send_message("Leave approved and confirmed by HR.", ephemeral=True)

                    # Send a message to the leave applier
                    user = await client.fetch_user(self.user_id)
                    await user.send("Your leave has been approved by the HR and has been confirmed.")

                elif self.current_stage == 'third':
                    # Update the leave status in the database
                    c1.execute(f'''
                        UPDATE {table_name}
                        SET leave_status = 'Accepted'
                        WHERE leave_id = ?
                    ''', (self.leave_details['leave_id'],))
                    conn1.commit()

                    # Update the dynamic_updates table
                    leave_reason = self.leave_details['leave_reason'].lower()
                    number_of_days_off = self.leave_details['number_of_days_off']
                    if leave_reason == "sick":
                        c2.execute(f'''
                            UPDATE dynamic_updates
                            SET total_sick_leave = total_sick_leave + ?, last_leave_taken = ?
                            WHERE user_id = ?
                        ''', (number_of_days_off, self.leave_details['date_to'], self.user_id))
                    elif leave_reason == "casual":
                        c2.execute(f'''
                            UPDATE dynamic_updates
                            SET total_casual_leave = total_casual_leave + ?, last_leave_taken = ?
                            WHERE user_id = ?
                        ''', (number_of_days_off, self.leave_details['date_to'], self.user_id))
                    elif leave_reason == "c. off":
                        c2.execute(f'''
                            UPDATE dynamic_updates
                            SET total_c_off = total_c_off + ?, last_leave_taken = ?
                            WHERE user_id = ?
                        ''', (number_of_days_off, self.leave_details['date_to'], self.user_id))
                    conn2.commit()

                    # Update the footer text in the database
                    new_footer_text = f"Stage: final | User ID: {self.user_id} | Nickname: {self.nickname} | Channel ID: {approval_channels['pa']} | Message ID: {interaction.message.id}"
                    embed.set_footer(text=new_footer_text)
                    await interaction.message.edit(embed=embed)
                    c1.execute(f'''
                        UPDATE {table_name}
                        SET footer_text = ?
                        WHERE leave_id = ?
                    ''', (new_footer_text, self.leave_details['leave_id']))
                    conn1.commit()

                    # Update the original message to show it has been approved
                    await interaction.message.edit(view=None)
                    await interaction.message.edit(content="Leave Approved", view=View().add_item(Button(label="Approved", style=discord.ButtonStyle.success, disabled=True)))
                    await interaction.response.send_message("Leave approved and confirmed by the Principal Architect.", ephemeral=True)

                    # Send a message to the leave applier
                    user = await client.fetch_user(self.user_id)
                    await user.send("Your leave has been approved by the Principal Architect and has been confirmed.")

            else:
                # Handle decline logic
                await interaction.response.send_modal(DeclineReasonModal(self.user_id, self.leave_details, self.current_stage, self.nickname))
        except Exception as e:
            print(f"Error in handle_approval: {str(e)}")
            try:
                await interaction.followup.send("An error occurred while processing the approval.", ephemeral=True)
            except discord.errors.NotFound:
                pass

class DeclineReasonModal(Modal):
    def __init__(self, user_id, leave_details, current_stage, nickname):
        super().__init__(title="Reason for Decline")
        self.user_id = user_id
        self.leave_details = leave_details
        self.current_stage = current_stage
        self.nickname = nickname  # Store the name of the leave applier
        self.add_item(TextInput(label="Reason for Decline", style=discord.TextStyle.short, max_length=1024, placeholder="Enter the reason for decline"))

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer()  # Acknowledge the interaction
        try:
            reason_for_decline = self.children[0].value.strip()
            self.leave_details['reason_for_decline'] = reason_for_decline

            # Create the embed with the current stage
            embed = create_leave_embed(self.leave_details, self.user_id, self.nickname, self.current_stage)
            embed.add_field(name="Declined By", value=interaction.user.display_name, inline=False)
            embed.add_field(name="Reason for Decline", value=reason_for_decline, inline=False)

            if self.current_stage == 'first':
                # Update the original message to show it has been declined
                await interaction.message.edit(view=None)
                await interaction.message.edit(content="Leave Declined", view=View().add_item(Button(label="Declined", style=discord.ButtonStyle.danger, disabled=True)))
                await interaction.followup.send("Leave declined and sent to the appropriate channel.", ephemeral=True)

                # Send a message to the leave applier
                user = await client.fetch_user(self.user_id)
                await user.send(f"Your leave has been declined by your head of department. Reason: {reason_for_decline}. Please contact HR for further assistance.")

            elif self.current_stage == 'second':
                # Update the original message to show it has been declined
                await interaction.message.edit(view=None)
                await interaction.message.edit(content="Leave Declined", view=View().add_item(Button(label="Declined", style=discord.ButtonStyle.danger, disabled=True)))
                await interaction.followup.send("Leave declined and sent to the appropriate channel.", ephemeral=True)

                # Send a message to the leave applier
                user = await client.fetch_user(self.user_id)
                await user.send(f"Your leave has been declined by HR. Reason: {reason_for_decline}.")

                # Send the embed to the PA channel for third approval
                pa_channel = client.get_channel(approval_channels['pa'])
                if pa_channel:
                    await pa_channel.send(embed=embed, view=LeaveApprovalView(self.user_id, self.leave_details, 'third', self.nickname))
                else:
                    raise ValueError(f"Channel with ID {approval_channels['pa']} not found.")

            elif self.current_stage == 'third':
                # Update the original message to show it has been declined
                await interaction.message.edit(view=None)
                await interaction.message.edit(content="Leave Declined", view=View().add_item(Button(label="Declined", style=discord.ButtonStyle.danger, disabled=True)))
                await interaction.followup.send("Leave declined and sent to the appropriate channel.", ephemeral=True)

                # Send a message to the leave applier
                user = await client.fetch_user(self.user_id)
                await user.send(f"Your leave has been declined by the Principal Architect. Reason: {reason_for_decline}.")
        except Exception as e:
            print(f"Error in DeclineReasonModal on_submit: {str(e)}")
            try:
                await interaction.followup.send("An error occurred while processing the decline.", ephemeral=True)
            except discord.errors.NotFound:
                pass

class WithdrawLeaveModal(Modal):
    def __init__(self):
        super().__init__(title="Withdraw Leave Application")
        self.add_item(TextInput(label="Leave ID", style=discord.TextStyle.short, max_length=10, placeholder="Enter Leave ID"))

    async def on_submit(self, interaction: discord.Interaction):
        try:
            leave_id = int(self.children[0].value.strip())
            user_id = interaction.user.id
            nickname = interaction.user.display_name
            table_name = sanitize_table_name(nickname)
            
            # Check if the leave ID exists and is accepted
            result = await db_execute(c1.execute, f'''
                SELECT leave_reason, number_of_days_off FROM {table_name} WHERE leave_id = ? AND leave_status = 'Accepted'
            ''', (leave_id,))
            result = result.fetchone()
            
            if result:
                leave_reason, number_of_days_off = result
                # Verify that the user requesting the withdrawal is the same as the user who applied for the leave
                dynamic_result = await db_execute(c2.execute, '''
                    SELECT user_id FROM dynamic_updates WHERE nickname = ?
                ''', (nickname,))
                dynamic_result = dynamic_result.fetchone()
                if dynamic_result and dynamic_result[0] != user_id:
                    await interaction.response.send_message("You can only withdraw your own leave applications.", ephemeral=True)
                    return
                
                # Withdraw the leave
                await db_execute(c1.execute, f'''
                    UPDATE {table_name} SET leave_status = ? WHERE leave_id = ?
                ''', ("Withdrawn", leave_id))
                await db_execute(conn1.commit)
                
                # Update the dynamic_updates table
                leave_reason = leave_reason.lower()
                if leave_reason == "sick":
                    await db_execute(c2.execute, '''
                        UPDATE dynamic_updates
                        SET total_sick_leave = total_sick_leave - ?
                        WHERE user_id = ?
                    ''', (number_of_days_off, user_id))
                elif leave_reason == "casual":
                    await db_execute(c2.execute, '''
                        UPDATE dynamic_updates
                        SET total_casual_leave = total_casual_leave - ?
                        WHERE user_id = ?
                    ''', (number_of_days_off, user_id))
                elif leave_reason == "c. off":
                    await db_execute(c2.execute, '''
                        UPDATE dynamic_updates
                        SET total_c_off = total_c_off - ?
                        WHERE user_id = ?
                    ''', (number_of_days_off, user_id))
                await db_execute(conn2.commit)

                # Update the last_leave_taken field with the latest accepted leave date
                latest_leave_date = await db_execute(c1.execute, f'''
                    SELECT MAX(date_to) FROM {table_name} WHERE leave_status = 'Accepted'
                ''')
                latest_leave_date = latest_leave_date.fetchone()[0]
                if latest_leave_date:
                    await db_execute(c2.execute, '''
                        UPDATE dynamic_updates
                        SET last_leave_taken = ?
                        WHERE user_id = ?
                    ''', (latest_leave_date, user_id))
                    await db_execute(conn2.commit)

                await interaction.response.send_message(f"Leave with ID {leave_id} has been withdrawn.", ephemeral=True)
            else:
                await interaction.response.send_message(f"Leave with ID {leave_id} not found or not accepted.", ephemeral=True)
        except ValueError:
            await interaction.response.send_message("Invalid Leave ID. Please enter a valid number.", ephemeral=True)
        except Exception as e:
            await interaction.response.send_message(f"An error occurred: {e}", ephemeral=True)

async def send_leave_application_to_approval_channel(interaction, leave_details, user_roles):
    # Determine the first approval channel based on user roles
    first_approval_channel_id = None
    second_approval_channel_id = approval_channels['hr']  # Default second approval channel

    for role_name, role_id in department_roles.items():
        if role_id in user_roles:
            first_approval_channel_id = approval_channels[role_name]
            break

    # Check if the user has a role that requires direct second approval
    for role_name, role_id in direct_second_approval_roles.items():
        if role_id in user_roles:
            first_approval_channel_id = None  # Skip the first approval channel
            break

    # Determine the appropriate approval channel
    approval_channel_id = first_approval_channel_id if first_approval_channel_id else second_approval_channel_id

    channel = client.get_channel(approval_channel_id)
    if channel is None:
        await interaction.response.send_message(f"Error: Could not find the approval channel with ID {approval_channel_id}. Please check the channel configuration.", ephemeral=True)
        print(f"Available channels: {[channel.id for channel in interaction.guild.channels]}")
        return

    current_stage = 'first' if first_approval_channel_id else 'hr'
    view = LeaveApprovalView(interaction.user.id, leave_details, current_stage, interaction.user.display_name)
    embed = create_leave_embed(leave_details, interaction.user.id, interaction.user.display_name, current_stage)
    message = await channel.send(embed=embed, view=view)

    # Include the message ID in the footer text
    footer_text = f"Stage: {current_stage} | User ID: {interaction.user.id} | Nickname: {interaction.user.display_name} | Channel ID: {approval_channel_id} | Message ID: {message.id}"
    embed.set_footer(text=footer_text)
    await message.edit(embed=embed)

    # Store the footer text in the database
    table_name = sanitize_table_name(interaction.user.display_name)
    await db_execute(c1.execute, f'''
        UPDATE {table_name}
        SET footer_text = ?
        WHERE leave_id = ?
    ''', (footer_text, leave_details['leave_id']))
    await db_execute(conn1.commit)

class FullDayLeaveModal(Modal):
    def __init__(self):
        super().__init__(title="FULL DAY LEAVE APPLICATION")
        self.add_item(TextInput(label="LEAVE REASON", style=discord.TextStyle.short, max_length=10, placeholder="CASUAL / SICK / C. OFF"))
        self.add_item(TextInput(label="DATE - FROM DD-MM-YYYY TO DD-MM-YYYY", style=discord.TextStyle.short, max_length=1024, placeholder="DD-MM-YYYY TO DD-MM-YYYY"))
        self.add_item(TextInput(label="NO. OF DAYS OFF", style=discord.TextStyle.short, max_length=1024, placeholder="NUMBER OF DAYS OFF"))
        self.add_item(TextInput(label="RESUME OFFICE ON - DD-MM-YYYY", style=discord.TextStyle.short, max_length=1024, placeholder="DD-MM-YYYY"))

    async def on_submit(self, interaction: discord.Interaction):
        try:
            user_id = interaction.user.id
            nickname = interaction.user.display_name
            table_name = sanitize_table_name(nickname)
            
            # Convert inputs to uppercase
            leave_reason = self.children[0].value.strip().upper()
            date_range = self.children[1].value.strip().upper().split(" TO ")
            number_of_days_off = self.children[2].value.strip().upper()
            resume_office_on = self.children[3].value.strip().upper()
            
            # Validate leave reason
            if leave_reason not in ["CASUAL", "SICK", "C. OFF"]:
                await interaction.response.send_message("INVALID LEAVE REASON. PLEASE ENTER 'CASUAL', 'SICK', OR 'C. OFF'.", ephemeral=True)
                return
            
            # Validate date range
            if len(date_range) != 2:
                await interaction.response.send_message("INVALID DATE RANGE FORMAT. PLEASE USE 'DD-MM-YYYY TO DD-MM-YYYY'.", ephemeral=True)
                return
            date_from, date_to = date_range
            try:
                datetime.strptime(date_from, "%d-%m-%Y")
                datetime.strptime(date_to, "%d-%m-%Y")
            except ValueError:
                await interaction.response.send_message("INVALID DATE FORMAT. PLEASE USE 'DD-MM-YYYY'.", ephemeral=True)
                return
            
            # Validate number of days off
            try:
                number_of_days_off = float(number_of_days_off)
            except ValueError:
                await interaction.response.send_message("INVALID INPUT FOR NUMBER OF DAYS OFF. PLEASE ENTER A VALID NUMBER.", ephemeral=True)
                return
            
            # Validate resume office on date
            try:
                datetime.strptime(resume_office_on, "%d-%m-%Y")
            except ValueError:
                await interaction.response.send_message("INVALID DATE FORMAT FOR RESUME OFFICE ON. PLEASE USE 'DD-MM-YYYY'.", ephemeral=True)
                return
            
            leave_details = {
                'leave_type': 'FULL DAY',
                'leave_reason': leave_reason,
                'date_from': date_from,
                'date_to': date_to,
                'number_of_days_off': number_of_days_off,
                'resume_office_on': resume_office_on
            }
            data = (
                leave_details['leave_type'],
                leave_details['leave_reason'],
                leave_details['date_from'],
                leave_details['date_to'],
                leave_details['number_of_days_off'],
                leave_details['resume_office_on'],
                None,  # time_off is None for full day leave
                "PENDING",  # leave_status is initially "Pending"
                None  # reason_for_decline is None initially
            )

            # Retry mechanism for database operations
            retries = 5
            for attempt in range(retries):
                try:
                    c1.execute(f'''
                        INSERT INTO {table_name} (leave_type, leave_reason, date_from, date_to, number_of_days_off, resume_office_on, time_off, leave_status, reason_for_decline)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', data)
                    conn1.commit()
                    break
                except sqlite3.OperationalError as e:
                    if "database is locked" in str(e) and attempt < retries - 1:
                        time.sleep(1)  # Wait for 1 second before retrying
                    else:
                        raise

            leave_id = c1.lastrowid
            leave_details['leave_id'] = leave_id  # Include leave_id in leave_details

            await send_leave_application_to_approval_channel(interaction, leave_details, [role.id for role in interaction.user.roles])
            await interaction.response.send_message(f"YOUR FULL DAY LEAVE APPLICATION HAS BEEN SUBMITTED. LEAVE ID: {leave_id}", ephemeral=True)
        except Exception as e:
            if not interaction.response.is_done():
                await interaction.response.send_message(f"AN ERROR OCCURRED: {e}", ephemeral=True)
            else:
                await interaction.followup.send(f"AN ERROR OCCURRED: {e}", ephemeral=True)

class HalfDayLeaveModal(Modal):
    def __init__(self):
        super().__init__(title="HALF DAY LEAVE APPLICATION")
        self.add_item(TextInput(label="LEAVE REASON", style=discord.TextStyle.short, max_length=10, placeholder="CASUAL / SICK / C. OFF"))
        self.add_item(TextInput(label="DATE", style=discord.TextStyle.short, max_length=1024, placeholder="DD-MM-YYYY"))
        self.add_item(TextInput(label="TIME PERIOD", style=discord.TextStyle.short, max_length=1024, placeholder="FORENOON/AFTERNOON"))

    async def on_submit(self, interaction: discord.Interaction):
        try:
            user_id = interaction.user.id
            nickname = interaction.user.display_name
            table_name = sanitize_table_name(nickname)
            
            # Convert inputs to uppercase
            leave_reason = self.children[0].value.strip().upper()
            date = self.children[1].value.strip().upper()
            time_period = self.children[2].value.strip().upper()
            
            # Validate leave reason
            if leave_reason not in ["CASUAL", "SICK", "C. OFF"]:
                await interaction.response.send_message("INVALID LEAVE REASON. PLEASE ENTER 'CASUAL', 'SICK', OR 'C. OFF'.", ephemeral=True)
                return
            
            # Validate date
            try:
                datetime.strptime(date, "%d-%m-%Y")
            except ValueError:
                await interaction.response.send_message("INVALID DATE FORMAT. PLEASE USE 'DD-MM-YYYY'.", ephemeral=True)
                return
            
            # Validate time period
            if time_period not in ["FORENOON", "AFTERNOON"]:
                await interaction.response.send_message("INVALID TIME PERIOD. PLEASE ENTER 'FORENOON' OR 'AFTERNOON'.", ephemeral=True)
                return
            
            leave_details = {
                'leave_type': 'HALF DAY',
                'leave_reason': leave_reason,
                'date_from': date,
                'date_to': None,  # date_to is None for half day leave
                'number_of_days_off': 0.5,  # number_of_days_off is 0.5 for half day leave
                'resume_office_on': None,  # resume_office_on is None for half day leave
                'time_period': time_period
            }
            data = (
                leave_details['leave_type'],
                leave_details['leave_reason'],
                leave_details['date_from'],
                leave_details['date_to'],
                leave_details['number_of_days_off'],
                leave_details['resume_office_on'],
                leave_details['time_period'],
                "PENDING",  # leave_status is initially "Pending"
                None  # reason_for_decline is None initially
            )

            # Retry mechanism for database operations
            retries = 5
            for attempt in range(retries):
                try:
                    c1.execute(f'''
                        INSERT INTO {table_name} (leave_type, leave_reason, date_from, date_to, number_of_days_off, resume_office_on, time_period, leave_status, reason_for_decline)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', data)
                    conn1.commit()
                    break
                except sqlite3.OperationalError as e:
                    if "database is locked" in str(e) and attempt < retries - 1:
                        time.sleep(1)  # Wait for 1 second before retrying
                    else:
                        raise

            leave_id = c1.lastrowid
            leave_details['leave_id'] = leave_id  # Include leave_id in leave_details

            await send_leave_application_to_approval_channel(interaction, leave_details, [role.id for role in interaction.user.roles])
            await interaction.response.send_message(f"YOUR HALF DAY LEAVE APPLICATION HAS BEEN SUBMITTED. LEAVE ID: {leave_id}", ephemeral=True)
        except Exception as e:
            if not interaction.response.is_done():
                await interaction.response.send_message(f"AN ERROR OCCURRED: {e}", ephemeral=True)
            else:
                await interaction.followup.send(f"AN ERROR OCCURRED: {e}", ephemeral=True)

class OffDutyLeaveModal(Modal):
    def __init__(self):
        super().__init__(title="OFF DUTY APPLICATION")
        self.add_item(TextInput(label="LEAVE REASON", style=discord.TextStyle.short, max_length=10, placeholder="CASUAL / SICK / C. OFF"))
        self.add_item(TextInput(label="DATE", style=discord.TextStyle.short, max_length=1024, placeholder="DD-MM-YYYY"))
        self.add_item(TextInput(label="TIME OFF", style=discord.TextStyle.short, max_length=1024, placeholder="HH-MM AM/PM TO HH-MM AM/PM"))
        self.add_item(TextInput(label="CUMULATED HOURS", style=discord.TextStyle.short, max_length=1024, placeholder="NO. OF HOURS"))

    async def on_submit(self, interaction: discord.Interaction):
        try:
            user_id = interaction.user.id
            nickname = interaction.user.display_name
            table_name = sanitize_table_name(nickname)
            
            # Convert inputs to uppercase
            leave_reason = self.children[0].value.strip().upper()
            date = self.children[1].value.strip().upper()
            time_off = self.children[2].value.strip().upper()
            cumulated_hours_str = self.children[3].value.strip().upper()
            
            # Validate leave reason
            if leave_reason not in ["CASUAL", "SICK", "C. OFF"]:
                await interaction.response.send_message("INVALID LEAVE REASON. PLEASE ENTER 'CASUAL', 'SICK', OR 'C. OFF'.", ephemeral=True)
                return
            
            # Validate date
            try:
                datetime.strptime(date, "%d-%m-%Y")
            except ValueError:
                await interaction.response.send_message("INVALID DATE FORMAT. PLEASE USE 'DD-MM-YYYY'.", ephemeral=True)
                return
            
            # Validate time off
            if not re.match(r'^\d{2}-\d{2} (AM|PM) TO \d{2}-\d{2} (AM|PM)$', time_off):
                await interaction.response.send_message("INVALID TIME OFF FORMAT. PLEASE USE 'HH-MM AM/PM TO HH-MM AM/PM'.", ephemeral=True)
                return
            
            # Validate cumulated hours
            try:
                cumulated_hours = float(re.findall(r'\d+\.?\d*', cumulated_hours_str)[0])  # Extract the numeric part
            except (ValueError, IndexError):
                await interaction.response.send_message("INVALID INPUT FOR CUMULATED HOURS. PLEASE ENTER A VALID NUMBER.", ephemeral=True)
                return
            
            leave_details = {
                'leave_type': "OFF DUTY",
                'leave_reason': leave_reason,
                'date_from': date,  # date_from is the date for off duty leave
                'date_to': None,  # date_to is None for off duty leave
                'number_of_days_off': None,  # number_of_days_off is None for off duty leave
                'resume_office_on': None,  # resume_office_on is None for off duty leave
                'time_off': time_off
            }
            data = (
                leave_details['leave_type'],
                leave_details['leave_reason'],
                leave_details['date_from'],
                leave_details['date_to'],
                leave_details['number_of_days_off'],
                leave_details['resume_office_on'],
                None,  # time_period is None for off duty leave
                leave_details['time_off'],
                "PENDING",  # leave_status is initially "Pending"
                None  # reason_for_decline is None initially
            )

            # Retry mechanism for database operations
            retries = 5
            for attempt in range(retries):
                try:
                    c1.execute(f'''
                        INSERT INTO {table_name} (leave_type, leave_reason, date_from, date_to, number_of_days_off, resume_office_on, time_period, time_off, leave_status, reason_for_decline)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', data)
                    conn1.commit()
                    break
                except sqlite3.OperationalError as e:
                    if "database is locked" in str(e) and attempt < retries - 1:
                        time.sleep(1)  # Wait for 1 second before retrying
                    else:
                        raise

            leave_id = c1.lastrowid
            leave_details['leave_id'] = leave_id  # Include leave_id in leave_details

            # Update the off_duty_hours in the dynamic_updates table
            c2.execute('''
                UPDATE dynamic_updates
                SET off_duty_hours = off_duty_hours + ?
                WHERE user_id = ?
            ''', (cumulated_hours, user_id))
            conn2.commit()

            await send_leave_application_to_approval_channel(interaction, leave_details, [role.id for role in interaction.user.roles])
            await interaction.response.send_message(f"YOUR OFF DUTY APPLICATION HAS BEEN SUBMITTED. LEAVE ID: {leave_id}", ephemeral=True)
        except Exception as e:
            if not interaction.response.is_done():
                await interaction.response.send_message(f"AN ERROR OCCURRED: {e}", ephemeral=True)
            else:
                await interaction.followup.send(f"AN ERROR OCCURRED: {e}", ephemeral=True)

# Command to manually trigger the export function for testing
@client.command()
async def export_leave(ctx):
    try:
        # Specify the directory where you want to save the Excel file   
        export_directory = r"Z:\Bot databases\Leave details"
        os.makedirs(export_directory, exist_ok=True)  # Create the directory if it doesn't exist

        # Generate the file name with the current date
        current_date = datetime.now().strftime("%d-%m-%Y")
        export_path = os.path.join(export_directory, f'leave_details_export_{current_date}.xlsx')

        # Calculate the start and end dates of the current month
        start_of_month = datetime.now().replace(day=1).strftime("%d-%m-%Y")
        end_of_month = (datetime.now().replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        end_of_month = end_of_month.strftime("%d-%m-%Y")

        # Fetch all tables in the leave_details database, excluding sqlite_sequence
        c1.execute("SELECT name FROM sqlite_master WHERE type='table' AND name != 'sqlite_sequence';")
        tables = c1.fetchall()

        # Create a new workbook and remove the default sheet
        wb = Workbook()
        wb.remove(wb.active)

        for table_name in tables:
            table_name = table_name[0]
            # Truncate the table name to 31 characters for the sheet name
            sheet_name = table_name[:31]
            # Read the table into a DataFrame, including only accepted and withdrawn leaves within the month
            query = f"""
                SELECT * FROM {table_name}
                WHERE (leave_status = 'Accepted' OR leave_status = 'Withdrawn')
                AND (date_from >= '{start_of_month}' AND (date_to <= '{end_of_month}' OR date_to IS NULL))
            """
            df = pd.read_sql_query(query, conn1)
            
            if df.empty:
                continue  # Skip empty tables
            
            # Add a new sheet to the workbook
            ws = wb.create_sheet(title=sheet_name)
            
            # Write the DataFrame to the sheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Apply header formatting
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Create a table for better layout
            tab = Table(displayName=f"Table_{sheet_name}", ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # Check if any sheets were added to the workbook
        if len(wb.sheetnames) == 0:
            await ctx.send("No data available to export.")
            return

        # Save the workbook
        wb.save(export_path)
        await ctx.send(f"Leave details exported to {export_path}")
    except Exception as e:
        await ctx.send(f"An error occurred while exporting leave details: {str(e)}")

client.run(TOKEN)
