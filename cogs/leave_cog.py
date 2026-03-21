"""
cogs/leave_cog.py — Leave Management Cog
Copyright (c) 2026 Concord Desk. All rights reserved.
PROPRIETARY AND CONFIDENTIAL.
"""

import asyncio
import os
import logging
from datetime import datetime, timedelta

import discord
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from Bots.utils.timezone import IST
from Bots.db_managers import leave_db_manager as db

import cogs.leave_config as cfg
from cogs.leave_config import resolve_leave_config
from cogs.leave_views import (
    LeaveApplicationView,
    LeaveApprovalView,
    extract_leave_details_from_embed,
)

logger = logging.getLogger("Concord")


class LeaveCog(commands.Cog, name="Leave"):
    """Handles leave applications and approvals."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    async def cog_load(self):
        """Start the DB worker and initialize."""
        self.bot.loop.create_task(db.db_worker())
        await db.initialize_leave_db()
        # Ensure the discovery event exists on the bot
        if not hasattr(self.bot, 'discovery_complete'):
            self.bot.discovery_complete = asyncio.Event()

    @commands.Cog.listener()
    async def on_ready(self):
        # Wait for DiscoveryCog to finish its initial sweep
        logger.info("[Leave] Waiting for DiscoveryCog to complete initial sweep...")
        await self.bot.discovery_complete.wait()
        await resolve_leave_config()

        guild = self.bot.guilds[0] if self.bot.guilds else None
        if not guild:
            logger.warning("[ERR-LV-036] [Leave] No guilds found. Skipping member sync.")
            return

        emp_role = guild.get_role(cfg.EMP_ROLE_ID)
        await db.create_dynamic_table()

        if emp_role:
            for member in emp_role.members:
                await db.insert_dynamic_user(member.display_name, member.id)
        else:
            logger.warning(f"[ERR-LV-037] [Leave] Employee role (ID: {cfg.EMP_ROLE_ID}) not found in server. Skipping bulk member sync.")

        # Create base styled embed for Leave app logic
        leave_embed = discord.Embed(
            title="🌴 Leave Management System",
            description="Welcome to the Concord Leave Center. Please select your leave type using the buttons below.",
            color=0x2ecc71
        )
        leave_embed.add_field(name="Full Day", value="Apply for a standard full day leave.", inline=True)
        leave_embed.add_field(name="Half Day", value="Apply for a forenoon or afternoon half-day leave.", inline=True)
        leave_embed.add_field(name="Off Duty", value="Apply for hourly off-duty time.", inline=True)
        leave_embed.add_field(name="Leave Details", value="View your current leave balances.", inline=False)
        leave_embed.set_footer(text="Concord Unified Engine")

        # Register LeaveApplicationView globally so button interactions survive bot restarts
        self.bot.add_view(LeaveApplicationView())

        # Reattach submit channel buttons
        submit_channel = self.bot.get_channel(cfg.SUBMIT_CHANNEL_ID)
        if submit_channel:
            try:
                await submit_channel.set_permissions(
                    submit_channel.guild.default_role,
                    send_messages=False,
                    add_reactions=False,
                    create_public_threads=False,
                    create_private_threads=False,
                    send_messages_in_threads=False
                )
            except Exception as e:
                logger.warning(f"[Leave] Could not set permissions on leave-applications: {e}")

        found_existing = False
        async for message in submit_channel.history(limit=10):
            if (
                message.author == self.bot.user
                and message.embeds
                and ("Leave Management System" in message.embeds[0].title or "Please select your leave type" in message.embeds[0].title)
            ):
                view = LeaveApplicationView()
                await asyncio.sleep(1)
                await message.edit(embed=leave_embed, view=view)
                logger.info("[Leave] Reattached view to existing message in submit channel.")
                found_existing = True
                break
        if not found_existing:
            view = LeaveApplicationView()
            await asyncio.sleep(1)
            await submit_channel.send(embed=leave_embed, view=view)
            logger.info("[Leave] Created new leave application message in submit channel.")

        # Reattach approval channel buttons
        for channel_id in set(cfg.APPROVAL_CHANNELS.values()):
            channel = self.bot.get_channel(channel_id)
            if channel is None:
                continue
            async for message in channel.history(limit=50):
                if (
                    message.author == self.bot.user
                    and message.embeds
                    and "Leave Application" in message.embeds[0].title
                ):
                    if message.components and not any(btn.disabled for btn in message.components[0].children):
                        leave_details = extract_leave_details_from_embed(message.embeds[0])
                        footer_text = message.embeds[0].footer.text if message.embeds[0].footer else None
                        if footer_text:
                            parts = footer_text.split(" | ")
                            if len(parts) >= 5:
                                try:
                                    current_stage = parts[0].split(": ")[1]
                                    user_id = int(parts[1].split(": ")[1])
                                    nickname = parts[2].split(": ")[1]
                                    message_id = int(parts[4].split(": ")[1])
                                    view = LeaveApprovalView(user_id, leave_details, current_stage, nickname, bot_ref=self.bot)
                                    # Since we aren't in an interaction here, we need a slight shim for ensure_buttons
                                    class DummyInteraction: data = {}
                                    await view._ensure_buttons_attached(DummyInteraction())
                                    self.bot.add_view(view, message_id=message_id)
                                    logger.info(f"[Leave] Reattached approval view for message {message_id} in channel {channel_id}.")
                                except (IndexError, ValueError) as e:
                                    logger.warning(f"[Leave] Footer parse error in channel {channel_id}: {e}")

    @commands.Cog.listener()
    async def on_member_update(self, before: discord.Member, after: discord.Member):
        before_roles = [r.id for r in before.roles]
        after_roles = [r.id for r in after.roles]
        if cfg.EMP_ROLE_ID in after_roles and cfg.EMP_ROLE_ID not in before_roles:
            await db.insert_dynamic_user(after.display_name, after.id)
        elif cfg.EMP_ROLE_ID not in after_roles and cfg.EMP_ROLE_ID in before_roles:
            await db.remove_dynamic_user(after.id)

    @commands.command(name="export_leave")
    async def export_leave(self, ctx):
        """Exports current month's leave details to an Excel file."""
        try:
            export_directory = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Database', 'Leave details'))
            os.makedirs(export_directory, exist_ok=True)
            current_date = datetime.now(IST).strftime("%d-%m-%Y")
            export_path = os.path.join(export_directory, f"leave_details_export_{current_date}.xlsx")
            start_of_month = datetime.now(IST).replace(day=1).strftime("%d-%m-%Y")
            end_of_month = ((datetime.now(IST).replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)).strftime("%d-%m-%Y")

            tables = await db.get_all_tables()
            wb = Workbook()
            wb.remove(wb.active)

            for (table_name,) in tables:
                sheet_name = table_name[:31]
                df = await db.fetch_table_data(table_name, start_of_month, end_of_month)
                if df.empty:
                    continue

                # Make column headers human-readable (e.g. cancellation_reason -> Cancellation Reason)
                df.columns = [str(col).replace('_', ' ').title() for col in df.columns]

                ws = wb.create_sheet(title=sheet_name)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                tab = Table(
                    displayName=f"Table_{sheet_name}",
                    ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}",
                )
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
                ws.add_table(tab)
                for col in ws.columns:
                    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2

            if not wb.sheetnames:
                await ctx.send("No data available to export.")
                return
            wb.save(export_path)
            await ctx.send(f"Leave details exported to `{export_path}`")
        except Exception as e:
            await ctx.send(f"An error occurred: {e}")


async def setup(bot: commands.Bot):
    await bot.add_cog(LeaveCog(bot))
